# -*- coding: utf-8 -*-

import argparse
import json
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter
from datetime import datetime


def excel_to_json(input_file, sheet_name, start_row, end_row, start_col, end_col, output_file, transpose):
    """
    读取Excel文件并将指定范围的数据转换为JSON
      参数:
      input_file: 输入的Excel文件路径
      sheet_name: 要读取的工作表名称
      start_row: 起始行号(1 - based)
      end_row: 结束行号(1 - based)
      start_col: 起始列号(1 - based)
      end_col: 结束列号(1 - based)
      output_file: 输出的JSON文件路径
    """

    try:
        # 验证输入文件是否存在
        if not Path(input_file).is_file():
            raise FileNotFoundError(f"输入文件 '{input_file}' 不存在")

        # 加载工作簿
        # 如果不添加, data_only=True参数，则会读取到公式而不是最终值
        workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True)

        # 获取指定工作表
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在于文件中")

        sheet = workbook[sheet_name]

        # 验证行列范围
        if start_row < 1 or end_row < start_row:
            raise ValueError("行范围无效")
        if start_col < 1 or end_col < start_col:
            raise ValueError("列范围无效")

        # 读取数据
        excel_row_col_value = []
        # for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
        #                            min_col=start_col, max_col=end_col,
        #                            values_only=True):
        #     # 将每行数据转换为字典
        #     row_data = [cell_value for cell_value in row]
        #     excel_row_col_value.append(row_data)

        # 下方直接使用cell()方法访问每个单元格，这样可以更明确地控制每个单元格的读取方式
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # 获取单元格的值（如果是公式，将返回计算结果）
                cell_value = cell.value
                row_data.append(cell_value)
            excel_row_col_value.append(row_data)

        # 行列转置
        if transpose:
            excel_row_col_value = [list(row) for row in zip(*excel_row_col_value)]
            print("已执行行列转置")

        print(f"读取范围: {get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")

        # 转换为JSON
        json_data = json.dumps(excel_row_col_value, ensure_ascii=False, indent=2)

        # 保存到文件
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(json_data)

        print(f"成功将数据保存到 '{output_file}'")

        return excel_row_col_value

    except Exception as e:
        print(f"发生错误: {str(e)}")
    finally:
        if 'workbook' in locals():
            # noinspection PyUnboundLocalVariable
            workbook.close()


def parse_json_array(json_strings):
    """
    解析包含JSON字符串的数组，返回解析后的对象数组

    参数:
    json_strings -- 包含合法JSON字符串的列表（允许末尾有逗号）

    返回:
    包含解析后对象的列表
    """
    parsed_objects = []
    for json_str in json_strings:
        # 去除字符串两端的空白字符
        json_str = json_str.strip()

        # 检查末尾是否有逗号
        if json_str.endswith(','):
            json_str = json_str[:-1].strip()  # 删除逗号并再次去除可能的空白字符

        try:
            obj = json.loads(json_str)
            parsed_objects.append(obj)
        except json.JSONDecodeError as e:
            raise ValueError(f"输入的字符串不是合法的JSON: {json_str}") from e
    return parsed_objects


def update_json_property(json_file_path, property_path, new_value):
    """
    更新JSON文件中指定路径的属性值

    参数:
        json_file_path (str): JSON文件路径
        property_path (list): 表示多级属性名的字符串数组，例如 ['level1', 'level2', 'property']
        new_value (object): 要替换的新值

    返回:
        bool: 如果属性存在并成功更新返回True，否则返回False
    """

    try:
        # 1. 读取JSON文件
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)

        # 2. 检查属性路径是否存在
        current = data
        for i, key in enumerate(property_path[:-1]):
            if isinstance(current, list):
                current = current[int(key)]
                continue
            if key not in current:
                print(f"属性路径不存在: {'.'.join(property_path[:i + 1])}")
                return False
            current = current[key]

        last_key = property_path[-1]
        if last_key not in current:
            print(f"最终属性不存在: {'.'.join(property_path)}")
            return False

        # 3. 更新属性值
        current[last_key] = new_value

        # 4. 保存修改后的JSON文件
        with open(json_file_path, 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=2, ensure_ascii=False)

        return True

    except json.JSONDecodeError:
        print("错误: 文件不是有效的JSON格式")
        return False
    except Exception as e:
        print(f"发生错误: {str(e)}")
        return False


if __name__ == "__main__":
    # 设置命令行参数
    parser = argparse.ArgumentParser(description='将Excel文件中的指定范围数据转换为JSON')
    parser.add_argument('--start_row', type=int, default=3, help='起始行号(1-based)')
    parser.add_argument('--end_row', type=int, default=353, help='结束行号(1-based)')
    parser.add_argument('--start_col', type=int, default=13, help='起始列号(1-based)')
    parser.add_argument('--end_col', type=int, default=18, help='结束列号(1-based)')
    # parser.add_argument('--output', default=f'output_{datetime.now().strftime('%y%m%d_%H%M%S')}.json',
    #                     help='输出的JSON文件路径')

    args = parser.parse_args()
    xlsx_path = "/path/to/your_excel.xlsx"
    json_root = "/path/to/storage_json_directory"
    json_fields_d = ['data', 'items']
    json_fields_t = ['data', 'items', 0, 'data']
    for i in range(1, 4):
        sheet_name0 = "推广点位" + str(i)
        sheet_read_output_filename = f'sheet_{i}_{get_column_letter(args.start_col)}{args.start_row}:{get_column_letter(args.end_col)}{args.end_row}_{datetime.now().strftime('%y%m%d_%H%M%S')}.json'
        excel_read_results = excel_to_json(
            input_file=xlsx_path,
            sheet_name=sheet_name0,
            start_row=args.start_row,
            end_row=args.end_row,
            start_col=args.start_col,
            end_col=args.end_col,
            output_file=sheet_read_output_filename,
            transpose=True,
        )
        update_json_property(f"{json_root}/ios_{i}_detail.json", json_fields_d, parse_json_array(excel_read_results[1]))
        update_json_property(f"{json_root}/ios_{i}_pv_trend.json", json_fields_t,
                             parse_json_array(excel_read_results[3]))
        update_json_property(f"{json_root}/ios_{i}_uv_trend.json", json_fields_t,
                             parse_json_array(excel_read_results[5]))
        update_json_property(f"{json_root}/android_{i}_detail.json", json_fields_d, parse_json_array(excel_read_results[0]))
        update_json_property(f"{json_root}/android_{i}_pv_trend.json", json_fields_t,
                             parse_json_array(excel_read_results[2]))
        update_json_property(f"{json_root}/android_{i}_uv_trend.json", json_fields_t,
                             parse_json_array(excel_read_results[4]))
